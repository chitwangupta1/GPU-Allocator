"""
GPU Compute Sharing Marketplace - CrewAI Agent System with Gemini API
Python 3.10.10 Compatible - With Excel Integration
"""

from crewai import Agent, Task, Crew, Process, LLM
from typing import List, Dict
import json
from datetime import datetime
import os
from dotenv import load_dotenv
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import traceback

# Load environment variables
load_dotenv()

# Excel file path
EXCEL_FILE = 'gpu_marketplace_data.xlsx'

# Initialize Gemini LLM using CrewAI's LLM class
def get_gemini_llm():
    # Get API key from environment
    api_key = os.getenv('GOOGLE_API_KEY')
    if not api_key:
        raise ValueError("GOOGLE_API_KEY not found in environment variables. Please add it to your .env file.")
    
    # Using gemini-1.5-flash-002 for better rate limits on free tier
    # Alternative models available:
    # - gemini/gemini-2.0-flash-exp (experimental, lower rate limits)
    # - gemini/gemini-1.5-pro-002 (more capable but slower)
    return LLM(
        model="gemini/gemini-2.0-flash",
        api_key=api_key,
        temperature=0.7
    )

# Excel Database Manager
class ExcelDatabase:
    def __init__(self, filename=EXCEL_FILE):
        self.filename = filename
        self.init_excel()
    
    def init_excel(self):
        """Initialize Excel file with proper structure"""
        if not os.path.exists(self.filename):
            wb = Workbook()
            
            # Lenders Sheet
            ws_lenders = wb.active
            ws_lenders.title = "Lenders"
            headers_lenders = ['ID', 'Company Name', 'Email', 'GPU Type', 'Total GPUs', 
                             'Available GPUs', 'Memory (GB)', 'Location', 'Price/Hour', 
                             'Availability Period', 'Notes', 'Date Added']
            ws_lenders.append(headers_lenders)
            
            # Style headers
            for cell in ws_lenders[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Borrowers Sheet
            ws_borrowers = wb.create_sheet("Borrowers")
            headers_borrowers = ['ID', 'Company Name', 'Email', 'GPU Type', 'Quantity', 
                               'Duration', 'Workload', 'Budget/Hour', 'Location', 
                               'Requirements', 'Date Requested']
            ws_borrowers.append(headers_borrowers)
            
            for cell in ws_borrowers[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Allocations Sheet
            ws_allocations = wb.create_sheet("Allocations")
            headers_allocations = ['Allocation ID', 'Lender ID', 'Lender Company', 
                                  'Borrower ID', 'Borrower Company', 'GPU Type', 
                                  'Quantity Allocated', 'Start Date', 'End Date', 
                                  'Price/Hour', 'Status', 'Notes']
            ws_allocations.append(headers_allocations)
            
            for cell in ws_allocations[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            wb.save(self.filename)
            print(f"✓ Created Excel file: {self.filename}")
    
    def add_lender(self, data: Dict) -> int:
        """Add a new lender to Excel"""
        wb = load_workbook(self.filename)
        ws = wb['Lenders']
        
        lender_id = ws.max_row  # ID starts from 1 (excluding header)
        row = [
            lender_id,
            data.get('company_name'),
            data.get('email'),
            data.get('gpu_type'),
            data.get('quantity'),
            data.get('quantity'),  # Initially all GPUs are available
            data.get('memory_gb'),
            data.get('location'),
            data.get('price_per_hour'),
            data.get('availability'),
            data.get('notes', ''),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ]
        ws.append(row)
        wb.save(self.filename)
        print(f"✓ Added lender ID {lender_id} to Excel")
        return lender_id
    
    def add_borrower(self, data: Dict) -> int:
        """Add a new borrower to Excel"""
        wb = load_workbook(self.filename)
        ws = wb['Borrowers']
        
        borrower_id = ws.max_row if ws.max_row > 1 else 1

        row = [
            borrower_id,
            data.get('company_name'),
            data.get('email'),
            data.get('gpu_type'),
            data.get('quantity'),
            data.get('duration'),
            data.get('workload_type'),
            data.get('budget_per_hour'),
            data.get('location', ''),
            data.get('requirements', ''),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ]
        ws.append(row)
        wb.save(self.filename)
        print(f"✓ Added borrower ID {borrower_id} to Excel")
        return borrower_id
    
    def get_all_lenders(self) -> List[Dict]:
        """Get all lenders from Excel"""
        try:
            df = pd.read_excel(self.filename, sheet_name='Lenders')
            # Convert to dict and rename columns to match frontend expectations
            lenders = []
            for _, row in df.iterrows():
                lender = {
                    'id': int(row['ID']) if pd.notna(row['ID']) else None,
                    'company_name': str(row['Company Name']) if pd.notna(row['Company Name']) else '',
                    'email': str(row['Email']) if pd.notna(row['Email']) else '',
                    'gpu_type': str(row['GPU Type']) if pd.notna(row['GPU Type']) else '',
                    'quantity': int(row['Total GPUs']) if pd.notna(row['Total GPUs']) else 0,
                    'available_gpus': int(row['Available GPUs']) if pd.notna(row['Available GPUs']) else 0,
                    'memory_gb': int(row['Memory (GB)']) if pd.notna(row['Memory (GB)']) else 0,
                    'location': str(row['Location']) if pd.notna(row['Location']) else '',
                    'price_per_hour': float(row['Price/Hour']) if pd.notna(row['Price/Hour']) else 0.0,
                    'availability': str(row['Availability Period']) if pd.notna(row['Availability Period']) else '',
                    'notes': str(row['Notes']) if pd.notna(row['Notes']) else '',
                    'date_added': str(row['Date Added']) if pd.notna(row['Date Added']) else ''
                }
                lenders.append(lender)
            return lenders
        except Exception as e:
            print(f"Error reading lenders: {e}")
            traceback.print_exc()
            return []
    
    def get_available_lenders(self) -> List[Dict]:
        """Get lenders with available GPUs"""
        lenders = self.get_all_lenders()
        return [l for l in lenders if l.get('available_gpus', 0) > 0]
    
    def get_all_borrowers(self) -> List[Dict]:
        """Get all borrowers from Excel"""
        try:
            df = pd.read_excel(self.filename, sheet_name='Borrowers')
            # Convert to dict and rename columns to match frontend expectations
            borrowers = []
            for _, row in df.iterrows():
                borrower = {
                    'id': int(row['ID']) if pd.notna(row['ID']) else None,
                    'company_name': str(row['Company Name']) if pd.notna(row['Company Name']) else '',
                    'email': str(row['Email']) if pd.notna(row['Email']) else '',
                    'gpu_type': str(row['GPU Type']) if pd.notna(row['GPU Type']) else '',
                    'quantity': int(row['Quantity']) if pd.notna(row['Quantity']) else 0,
                    'duration': str(row['Duration']) if pd.notna(row['Duration']) else '',
                    'workload_type': str(row['Workload']) if pd.notna(row['Workload']) else '',
                    'budget_per_hour': float(row['Budget/Hour']) if pd.notna(row['Budget/Hour']) else 0.0,
                    'location': str(row['Location']) if pd.notna(row['Location']) else '',
                    'requirements': str(row['Requirements']) if pd.notna(row['Requirements']) else '',
                    'date_requested': str(row['Date Requested']) if pd.notna(row['Date Requested']) else ''
                }
                borrowers.append(borrower)
            return borrowers
        except Exception as e:
            print(f"Error reading borrowers: {e}")
            traceback.print_exc()
            return []
    
    def create_allocation(self, lender_id: int, borrower_id: int, quantity: int, 
                         duration: str, price: float, notes: str = '') -> int:
        """Create an allocation and update available GPUs"""
        wb = load_workbook(self.filename)
        
        # Get lender and borrower info
        ws_lenders = wb['Lenders']
        ws_borrowers = wb['Borrowers']
        
        lender_row = None
        lender_company = ''
        gpu_type = ''
        
        for row in ws_lenders.iter_rows(min_row=2, values_only=False):
            if row[0].value == lender_id:
                lender_row = row
                lender_company = row[1].value
                gpu_type = row[3].value
                available = row[5].value
                
                if available < quantity:
                    print(f"✗ Not enough GPUs available. Available: {available}, Requested: {quantity}")
                    return -1
                
                # Update available GPUs
                row[5].value = available - quantity
                break
        
        borrower_company = ''
        for row in ws_borrowers.iter_rows(min_row=2, values_only=True):
            if row[0] == borrower_id:
                borrower_company = row[1]
                break
        
        # Add allocation
        ws_allocations = wb['Allocations']
        allocation_id = ws_allocations.max_row
        
        allocation_row = [
            allocation_id,
            lender_id,
            lender_company,
            borrower_id,
            borrower_company,
            gpu_type,
            quantity,
            datetime.now().strftime('%Y-%m-%d'),
            f"{duration} from now",
            price,
            'Active',
            notes
        ]
        ws_allocations.append(allocation_row)
        
        wb.save(self.filename)
        print(f"✓ Created allocation ID {allocation_id}")
        return allocation_id
    
    def get_allocations(self) -> List[Dict]:
        """Get all allocations"""
        try:
            df = pd.read_excel(self.filename, sheet_name='Allocations')
            return df.to_dict('records')
        except Exception as e:
            print(f"Error reading allocations: {e}")
            return []

# Initialize database
db = ExcelDatabase()

# Initialize Gemini LLM instance with better error handling
gemini_llm = None
llm_error = None

try:
    gemini_llm = get_gemini_llm()
    print("✓ Gemini LLM initialized successfully")
    print(f"✓ Model: gemini-2.0-flash-exp")
except Exception as e:
    llm_error = str(e)
    print(f"✗ Warning: Could not initialize Gemini LLM")
    print(f"✗ Error: {e}")
    print(f"✗ Please check your GOOGLE_API_KEY in .env file")
    traceback.print_exc()

# Define Agents with Gemini
def create_agents():
    global gemini_llm, llm_error
    
    if gemini_llm is None:
        error_msg = f"Gemini LLM not initialized. "
        if llm_error:
            error_msg += f"Initialization error: {llm_error}. "
        error_msg += "Please ensure GOOGLE_API_KEY is set in your .env file and is valid."
        raise ValueError(error_msg)
    
    print("Creating AI agents...")
    
    gpu_analyst = Agent(
        role='GPU Market Analyst',
        goal='Analyze GPU requirements and availability to find optimal matches',
        backstory='''You are an expert in GPU computing resources with deep knowledge 
        of different GPU types (A100, H100, V100, etc.), their capabilities, and pricing. 
        You understand ML/AI workload requirements and can match them with appropriate hardware.''',
        verbose=True,
        allow_delegation=False,
        llm=gemini_llm
    )
    print("✓ GPU Analyst created")

    matching_specialist = Agent(
        role='Resource Matching Specialist',
        goal='Match GPU providers with requesters based on requirements and availability',
        backstory='''You specialize in matching computational resources with demand. 
        You consider factors like GPU type, memory, duration, and pricing 
        to create optimal matches between providers and requesters.''',
        verbose=True,
        allow_delegation=False,
        llm=gemini_llm
    )
    print("✓ Matching Specialist created")

    pricing_negotiator = Agent(
        role='Pricing and Contract Negotiator',
        goal='Determine fair pricing and contract terms for GPU sharing agreements',
        backstory='''You are an expert in cloud computing pricing models and contract 
        negotiation. You ensure fair pricing based on market rates, GPU specifications, 
        and usage duration while balancing the interests of both parties.''',
        verbose=True,
        allow_delegation=False,
        llm=gemini_llm
    )
    print("✓ Pricing Negotiator created")

    compliance_officer = Agent(
        role='Compliance and Security Officer',
        goal='Ensure all transactions meet security and compliance standards',
        backstory='''You verify that GPU sharing arrangements meet data security, 
        privacy regulations, and industry compliance standards. You assess risks 
        and ensure proper safeguards are in place.''',
        verbose=True,
        allow_delegation=False,
        llm=gemini_llm
    )
    print("✓ Compliance Officer created")
    
    return gpu_analyst, matching_specialist, pricing_negotiator, compliance_officer


class GPUMarketplaceCrew:
    def __init__(self):
        self.db = db
        self.agents = None
        self.initialization_error = None
        
        try:
            print("\n" + "="*60)
            print("Initializing AI Agents...")
            print("="*60)
            self.agents = create_agents()
            print("="*60)
            print("✓ All agents initialized successfully!")
            print("="*60 + "\n")
        except Exception as e:
            self.initialization_error = str(e)
            print("="*60)
            print(f"✗ Error creating agents: {e}")
            print("="*60 + "\n")
            traceback.print_exc()
    
    def rank_lenders(self, borrower_data, lenders):
        matches = []

        for l in lenders:
            score = 0

            # GPU match
            if borrower_data['gpu_type'] in l['gpu_type'] or borrower_data['gpu_type'] == "Any High-End":
                score += 40

            # Quantity
            if l['available_gpus'] >= borrower_data['quantity']:
                score += 30

            # Budget
            if l['price_per_hour'] <= borrower_data['budget_per_hour']:
                score += 20

            # Location (soft)
            if borrower_data.get('location') and borrower_data['location'] in l['location']:
                score += 10

            matches.append({**l, "score": score})

        return sorted(matches, key=lambda x: x["score"], reverse=True)[:3]

    
    def analyze_request(self, requester_data: Dict) -> Dict:
        """Analyze a GPU request and find matches from available lenders"""
        
        if self.agents is None:
            error_message = "Error: AI agents not initialized.\n\n"
            if self.initialization_error:
                error_message += f"Initialization Error: {self.initialization_error}\n\n"
            if llm_error:
                error_message += f"LLM Error: {llm_error}\n\n"
                error_message += "Troubleshooting steps:\n"
                error_message += "1. Check that GOOGLE_API_KEY is set in your .env file\n"
                error_message += "2. Verify your API key is valid (39 characters, starts with AIza)\n"
                error_message += "3. Test your key with: python test_env.py\n"
                error_message += "4. Make sure .env file is in the same directory as main.py\n"
                
                return {
                    "analysis_text": error_message,
                    "recommended_lenders": [],
                    "all_lenders": []
                }

        gpu_analyst, matching_specialist, pricing_negotiator, compliance_officer = self.agents
        
        try:
            # Get available lenders from Excel
            available_lenders = self.db.get_available_lenders()
            
            # Task 1: Analyze requirements
            analyze_task = Task(
                description=f'''Analyze this GPU compute request:
                Company: {requester_data.get('company_name')}
                GPU Type Needed: {requester_data.get('gpu_type')}
                Quantity: {requester_data.get('quantity')}
                Duration: {requester_data.get('duration')}
                Workload: {requester_data.get('workload_type')}
                Budget: ${requester_data.get('budget_per_hour')}/hour
                
                Provide a detailed analysis of the requirements and what to look for in providers.
                Focus on GPU specifications, performance needs, and compatibility.''',
                agent=gpu_analyst,
                expected_output='Detailed analysis of GPU requirements including recommendations'
            )
            
            # Task 2: Find matches from available lenders
            match_task = Task(
                description=f'''Based on the requirements analysis, review these AVAILABLE GPU lenders from our Excel database:
                {json.dumps(available_lenders, indent=2) if available_lenders else "No lenders with available GPUs currently"}
                
                Find the best matches considering:
                - GPU type compatibility
                - Available GPU quantity vs requested quantity ({requester_data.get('quantity')} GPUs needed)
                - Price comparison with budget (${requester_data.get('budget_per_hour')}/hour)
                
                Rank the top 3 matches with detailed reasoning. If no suitable lenders available, explain why.''',
                agent=matching_specialist,
                expected_output='Ranked list of matching lenders with detailed reasoning and availability status',
                context=[analyze_task]
            )
            
            # Task 3: Pricing analysis
            pricing_task = Task(
                description=f'''For the matched lenders, analyze pricing:
                - Compare lender prices with requester budget: ${requester_data.get('budget_per_hour')}/hour
                - Calculate total cost for duration: {requester_data.get('duration')}
                - Identify best value propositions
                - Suggest negotiation strategies if prices exceed budget''',
                agent=pricing_negotiator,
                expected_output='Comprehensive pricing analysis with cost calculations',
                context=[match_task]
            )
            
            # Task 4: Compliance check
            compliance_task = Task(
                description=f'''Review the proposed lender matches for:
                - Data security requirements for {requester_data.get('workload_type')} workloads
                - Geographic data residency concerns
                - Compliance standards (SOC2, GDPR, etc.)
                - Risk assessment for the allocation
                Provide final recommendations.''',
                agent=compliance_officer,
                expected_output='Detailed compliance assessment and final recommendations',
                context=[match_task]
            )
            
            # Create and run crew
            crew = Crew(
                agents=[gpu_analyst, matching_specialist, pricing_negotiator, compliance_officer],
                tasks=[analyze_task, match_task, pricing_task, compliance_task],
                process=Process.sequential,
                verbose=True
            )
            
            analysis_text = str(crew.kickoff())

            ranked_lenders = self.rank_lenders(
                requester_data,
                available_lenders
            )

            return {
                "analysis_text": analysis_text,
                "recommended_lenders": ranked_lenders,
                "all_lenders": available_lenders
            }

            
        except Exception as e:
            return {
                    "analysis_text": (
                        "Error during analysis.\n\n"
                        f"{str(e)}\n\n"
                        f"{traceback.format_exc()}"
                    ),
                    "recommended_lenders": [],
                    "all_lenders": []
                }
            
    def add_lender(self, lender_data: Dict) -> Dict:
        """Add a new GPU lender"""
        lender_id = self.db.add_lender(lender_data)
        return {
            'status': 'success',
            'message': f'Lender added successfully with ID: {lender_id}',
            'lender_id': lender_id
        }
    
    def add_borrower(self, borrower_data: Dict) -> Dict:
        """Add a new GPU borrower and find matches"""
        borrower_id = self.db.add_borrower(borrower_data)
        
        # Run crew analysis
        analysis_result = self.analyze_request(borrower_data)

        return {
            'status': 'success',
            'message': f'Request processed with ID: {borrower_id}',
            'borrower_id': borrower_id,
            'analysis': analysis_result['analysis_text'],
            'recommended_lenders': analysis_result['recommended_lenders']
        }

    
    def allocate_gpus(self, lender_id: int, borrower_id: int, quantity: int, 
                     duration: str, price: float, notes: str = '') -> Dict:
        """Allocate GPUs from lender to borrower"""
        allocation_id = self.db.create_allocation(lender_id, borrower_id, quantity, 
                                                   duration, price, notes)
        if allocation_id == -1:
            return {
                'status': 'error',
                'message': 'Not enough GPUs available or allocation failed'
            }
        
        return {
            'status': 'success',
            'message': f'Allocation created successfully with ID: {allocation_id}',
            'allocation_id': allocation_id
        }
    
    def get_all_lenders(self) -> List[Dict]:
        """Get all GPU lenders"""
        return self.db.get_all_lenders()
    
    def get_available_lenders(self) -> List[Dict]:
        """Get lenders with available GPUs"""
        return self.db.get_available_lenders()
    
    def get_all_borrowers(self) -> List[Dict]:
        """Get all GPU borrowers"""
        return self.db.get_all_borrowers()
    
    def get_allocations(self) -> List[Dict]:
        """Get all allocations"""
        return self.db.get_allocations()

# Flask API for web interface
from flask import Flask, render_template, request, jsonify
from flask_cors import CORS



app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}})

marketplace = GPUMarketplaceCrew()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/lenders', methods=['GET', 'POST'])
def lenders():
    if request.method == 'POST':
        data = request.json
        result = marketplace.add_lender(data)
        return jsonify(result)
    else:
        lenders_list = marketplace.get_all_lenders()
        print(f"Returning {len(lenders_list)} lenders")
        return jsonify({'lenders': lenders_list})

@app.route('/api/lenders/available', methods=['GET'])
def available_lenders():
    """Get only lenders with available GPUs"""
    lenders_list = marketplace.get_available_lenders()
    return jsonify({'lenders': lenders_list})

@app.route('/api/borrowers', methods=['GET', 'POST'])
def borrowers():
    if request.method == 'POST':
        data = request.json
        result = marketplace.add_borrower(data)
        return jsonify(result)
    else:
        borrowers_list = marketplace.get_all_borrowers()
        print(f"Returning {len(borrowers_list)} borrowers")
        return jsonify({'borrowers': borrowers_list})

@app.route('/api/allocations', methods=['GET', 'POST'])
def allocations():
    if request.method == 'POST':
        data = request.json
        result = marketplace.allocate_gpus(
            data.get('lender_id'),
            data.get('borrower_id'),
            data.get('quantity'),
            data.get('duration'),
            data.get('price'),
            data.get('notes', '')
        )
        return jsonify(result)
    else:
        allocations_list = marketplace.get_allocations()
        return jsonify({'allocations': allocations_list})

@app.route('/api/health')
def health():
    """Check if system is configured"""
    try:
        # Check if API key exists in environment (don't access or expose it)
        has_api_key = 'GOOGLE_API_KEY' in os.environ
        excel_exists = os.path.exists(EXCEL_FILE)
        
        if has_api_key and excel_exists:
            return jsonify({
                'status': 'healthy',
                'llm': 'Gemini 2.0 Flash',
                'configured': True,
                'excel_file': EXCEL_FILE,
                'excel_exists': excel_exists
            })
        else:
            return jsonify({
                'status': 'warning',
                'message': 'API key or Excel file missing',
                'configured': has_api_key,
                'excel_exists': excel_exists
            }), 500
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': 'Configuration error',
            'configured': False
        }), 500

if __name__ == "__main__":
    print("\n" + "=" * 60)
    print("GPU Compute Marketplace - Powered by Google Gemini")
    print("=" * 60)

    # Detect if running under Gunicorn
    running_under_gunicorn = "gunicorn" in os.environ.get("SERVER_SOFTWARE", "") \
        or "gunicorn" in sys.argv[0].lower()

    # Check API key
    if os.getenv("GOOGLE_API_KEY"):
        print("✓ Gemini API Key configured")
        print("✓ Using model: gemini-1.5-flash-002 (stable, better rate limits)")
    else:
        print("✗ WARNING: GOOGLE_API_KEY not found!")
        print("  Set it in Render → Environment Variables")

    # Check Excel file
    if os.path.exists(EXCEL_FILE):
        print(f"✓ Excel database found: {EXCEL_FILE}")
    else:
        print(f"✓ Creating new Excel database: {EXCEL_FILE}")

    # Only show localhost message + run dev server locally
    if not running_under_gunicorn:
        print("\nStarting LOCAL development server at http://localhost:5000")
        print("=" * 60 + "\n")
        app.run(host="0.0.0.0", port=5000, debug=True)
    else:
        print("\nRunning under Gunicorn (production mode)")
        print("Server binding handled by Gunicorn")
        print("=" * 60 + "\n")

